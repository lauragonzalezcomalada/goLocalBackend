from decimal import Decimal
from sqlite3 import IntegrityError
from django.shortcuts import render
from geopy.distance import geodesic
from collections import defaultdict,OrderedDict
import json
# Create your views here.
import mercadopago
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from .models import Activity, Bono, CampoReserva, EntradasForPlan, EventTemplate, ItemPlan, Payment, PaymentEventsRanges, Place, PrivatePlan, PrivatePlanInvitation, Promo, Reserva, ReservaForm, Tag, Ticket, UserProfile, User
from .serializers import ActivitySerializer, ActivitySerializerForGoLocalQR, BonoSerializer, CampoReservaSerializer, EventTemplateSerializer, ItemSerializer, PaymentEventsRangesSerializer, PlaceSerializer, PrivatePlanSerializer, PromoSerializer, ReservaFormSerializer, ReservaSerializer, TagSerializer, TicketSerializer, UserProfileBasicSerializer,UserProfileSerializer, UserProfileSerializerForQR
import ast 
from datetime import datetime, timedelta, date
from django.utils import timezone
from django.db.models import Count

from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework.exceptions import PermissionDenied

from django.core.mail import EmailMessage


@api_view(['GET'])
def hello_world(request):
    return Response({'message': 'Hello from Django!'})

from rest_framework.exceptions import AuthenticationFailed, PermissionDenied

class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):

    def validate(self, attrs):
        try:
            data = super().validate(attrs)
        except AuthenticationFailed:
            raise AuthenticationFailed("Email o contraseña incorrectos.")
        
        require_creador = self.context['request'].data.get('require_creador', False)
      
        if require_creador:
            user = self.user
            if not hasattr(user, 'profile') or not user.profile.creador:
                raise PermissionDenied("Este usuario no tiene permisos de creador") #codigo 403
        return data
    
class CustomTokenObtainPairView(TokenObtainPairView):
    print('hit el custom token view')
    serializer_class = CustomTokenObtainPairSerializer

    def post(self, request, *args, **kwargs):
        try:
            return super().post(request, *args, **kwargs)
        except AuthenticationFailed as e:
            return Response({'detail': str(e)}, status=401)
        except PermissionDenied as e:
            return Response({'detail': str(e)}, status=403)

@api_view(['GET'])
def get_all_places(request):
    # Obtenemos todos los objetos Place
    places = Place.objects.all()
    
    # Si no hay lugares en la base de datos
    if not places:
        return Response({'error': 'No places found'}, status=404)

    # Serializamos los objetos Place y los devolvemos en la respuesta
    serializer = PlaceSerializer(places, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def place_detail(request):

    name = request.query_params.get('name', None)
    if not name:
        return Response({'error': 'No name provided'}, status=400)
    
    try:
        place = Place.objects.get(name=name)
    except Place.DoesNotExist:
        return Response({'error': 'Place not found'}, status=404)

    serializer = PlaceSerializer(place)
    return Response(serializer.data)

@api_view(['GET'])
def activities(request): #hacer doble view
    
       # place_uuid = request.query_params.get('place_uuid',None)
    activity_uuid = request.query_params.get('activity_uuid',None)

      # if place_uuid:
      #      try:
      #          place = Place.objects.get(uuid=place_uuid)
      #      except Place.DoesNotExist:
      #          return Response({'error': 'No place found according to your uuid'}, status=404)
                    
      #      activities = Activity.objects.filter(place__uuid=place_uuid,startDateandtime__gte=timezone.now(),active=True).order_by('startDateandtime')
      #      print(activities)
            
       #     if not activities.exists:
       #         return Response({'error': 'No activities found according to the submitted place'}, status=404)

      #      serializer = ActivitySerializer(activities, many=True, fields=['uuid','name','shortDesc','place_name','image','startDateandtime','tag_detail','gratis','creador_image','asistentes'])
      #      return Response(serializer.data)
    if activity_uuid:
        try: 
            activity = Activity.objects.get(uuid=activity_uuid)
        except Activity.DoesNotExist:
            return Response({'error': 'No activity found according to your uuid'}, status=404)
        
        serializer = ActivitySerializer(activity,context={'request': request})
        return Response(serializer.data)
    else:
        activities = Activity.objects.filter(startDateandtime__gte=timezone.now(),active=True).order_by('startDateandtime')
        if not activities.exists:
            return Response({'error': 'No activities found'}, status=404)   
        serializer = ActivitySerializer(activities, many=True) #, fields=['uuid','name','shortDesc','place_name','image','startDateandtime','tag_detail','gratis','creador_image','asistentes'], context={'request': request})     
        return Response(serializer.data)
            #return Response({'error':'Neither place or activity submitted'},status=404)
    
        
    
@api_view(['GET'])
def promos(request):


    place_uuid = request.query_params.get('place_uuid',None)
    promo_uuid = request.query_params.get('promo_uuid',None)

    #if place_uuid:
    #    try:
    #        place = Place.objects.get(uuid=place_uuid)
    #    except Place.DoesNotExist:
    #        return Response({'error': 'No place found according to your uuid'}, status=404)
                
       
    if promo_uuid:
        try: 
            promo =  Promo.objects.get(uuid=promo_uuid)
        except Promo.DoesNotExist:
            return Response({'error': 'No promo found according to your uuid'}, status=404)
        
        serializer = PromoSerializer(promo, context={'request': request})
        return Response(serializer.data)
    else:
        promos = Promo.objects.filter(startDateandtime__gte=timezone.now(),active=True).order_by('startDateandtime')
        
        if not promos.exists:
            return Response({'error': 'No promos found according to the submitted place'}, status=404)

        serializer = PromoSerializer(promos, many=True, fields=['uuid','name','shortDesc','place_name','image','startDateandtime','endDateandtime','tag_detail','creador_image','asistentes'], context={'request': request})
        return Response(serializer.data)
      #  return Response({'error':'Neither place or activity submitted'},status=404)

@api_view(['GET'])
def registerView(request): 

    #GET ALL TAGS
    activity = request.GET.get('activity',None)
    promo = request.GET.get('promo',None)
    uuid = request.GET.get('uuid')
    
    if activity:
        try:
            activity = Activity.objects.get(uuid = uuid)
            activity.views +=1
            activity.save()
        except Activity.DoesNotExist:
            return Response('Error actualizando views de la actividad')
    elif promo: 
        try:
            promo = Promo.objects.get(uuid = uuid)
            promo.views +=1
            promo.save()
        except Promo.DoesNotExist:
            return Response('Error actualizando views de la actividad')
    return Response(status = 200)

@api_view(['GET'])
def registerShare(request): 

    #GET ALL TAGS
    activity = request.GET.get('activity',None)
    promo = request.GET.get('promo',None)
    uuid = request.GET.get('uuid')

    if activity:
        try:
            activity = Activity.objects.get(uuid = uuid)
            activity.shares +=1
            activity.save()
        except Activity.DoesNotExist:
            return Response('Error actualizando views de la actividad')
    elif promo: 
        try:
            promo = Promo.objects.get(uuid = uuid)
            promo.shares +=1
            promo.save()
        except Promo.DoesNotExist:
            return Response('Error actualizando views de la actividad')
    return Response(status = 200)
    

@api_view(['GET'])
def tags(request): 

    #GET ALL TAGS
    tags = Tag.objects.all()
    
    # Si no hay lugares en la base de datos
    if not tags:
        return Response({'error': 'No tags found'}, status=404)

    # Serializamos los objetos Place y los devolvemos en la respuesta
    serializer = TagSerializer(tags, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def validate_ticket(request):
    ticket_uuid = request.GET.get('ticket', None)
    event_uuid = request.GET.get('scanning_event', None)
    if not ticket_uuid:
        return Response({'error': 'Ticket UUID is required'}, status=400)

    try:
        ticket = Ticket.objects.get(uuid=ticket_uuid, entrada__activity__uuid=event_uuid)
    except Ticket.DoesNotExist:
        return Response({'error': 'Ticket not found, or not from this event'}, status=405)

    if ticket.status == 0: #No asistido, correcto, actualizamos a 1 y devolvemos success
        ticket.status = 1
        ticket.save()

        actividad = ticket.entrada.activity
        total_asistidos = Ticket.objects.filter(
            entrada__activity=actividad,
            status=1
        ).count()

        return Response({'asistidos': total_asistidos}, status = 200)
    elif ticket.status == 1: #Asistido, ticket ya usado
        return Response('El ticket ya fue usado antes', status= 404)


from django.core.mail import EmailMessage
from io import BytesIO
from reportlab.lib.pagesizes import A6
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph,Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import mm
from PIL import Image as PILImage




from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A6
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from django.utils import timezone
from io import BytesIO
from PIL import Image as PILImage

def generar_ticket_pdf(buyer_name, event_name, event_time, qr_code): #, background_path=None):

   # print('background path: ', background_path)
    buffer = BytesIO()
    width, height = A6
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A6,
        rightMargin=10*mm, leftMargin=10*mm,
        topMargin=10*mm, bottomMargin=10*mm
    )

    # --- estilos ---
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'title',
        parent=styles['Heading1'],
        fontSize=22,
        alignment=TA_CENTER,
        textColor='#FF6600'
    )
    normal_style = ParagraphStyle(
        'normal',
        parent=styles['Normal'],
        fontSize=18,
        alignment=TA_CENTER,
        textColor="#D34343"
    )

    # --- QR ---
    pil_img = PILImage.open(qr_code)
    qr_io = BytesIO()
    pil_img.save(qr_io, format='PNG')
    qr_io.seek(0)
    qr_img = Image(qr_io, width=60*mm, height=60*mm)
    qr_img.hAlign = 'CENTER'

    story = [
        Paragraph(event_name.upper(), title_style),
        Spacer(1, 5*mm),
        qr_img,
        Spacer(1, 5*mm),
        Paragraph(f"Comprador: {buyer_name}", normal_style),
        Spacer(1, 5*mm),
        Paragraph(f"Fecha: {timezone.localtime(event_time).strftime('%d/%m/%Y')}", normal_style),
        Spacer(1, 3*mm),
        Paragraph(f"Hora: {timezone.localtime(event_time).strftime('%H:%M')}", normal_style),
    ]

    # --- función para dibujar fondo ---
   # def fondo(canvas, doc):
      #  if background_path:
       #     bg = ImageReader(background_path)
       #     canvas.drawImage(bg, 0, 0, width=width, height=height)

    # Construir PDF con fondo en todas las páginas
    doc.build(story)#, onFirstPage=fondo, onLaterPages=fondo)

    buffer.seek(0)
    return buffer



from django.core.mail import send_mail
from django.conf import settings
@api_view(['POST'])
def create_ticket(request):
    data = request.data.copy()
    user = request.user

    if not user.is_authenticated:
        return Response({'error': 'User not authenticated'}, status=401)

    try:
        userProfile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response('User profile not found', status=404)

    entrada_uuid = data.get('entrada_uuid')
    if not entrada_uuid:
        return Response({'error': 'Entrada UUID is required'}, status=400)

    try:
        entrada = EntradasForPlan.objects.get(uuid=entrada_uuid)
    except EntradasForPlan.DoesNotExist:
        return Response({'error': 'Entrada not found'}, status=404)

    cantidad = data.get('cantidad')
    tickets = []

    for i in range(cantidad):  #map de ['tipus_entrada_1__uuid":{"amount..."}, 'tipus_entrada_2__uuid':{xxx}]
        if(entrada.disponibles > 0):
            ticket = Ticket.objects.create(
            user_profile=userProfile,
            entrada=entrada,
            nombre= data['name'],
            email= data['email'],
            fecha_compra=timezone.now(),
            precio = entrada.precio
            )
            entrada.disponibles = entrada.disponibles -1
            tickets.append(generar_ticket_pdf(data['name'],entrada.activity.name, entrada.activity.startDateandtime,ticket.qr_code)) #,os.path.join(settings.BASE_DIR, 'api', 'assets', 'backgroundticketsimage.png')))
        else:
            print('no quedan más entradas')


    entrada.save()
    userProfile.activities.add(entrada.activity)
    userProfile.save()


    email = EmailMessage(
        subject="Tu ticket para el evento 🎟️",
        body=f"Hola {data['name']}, adjuntamos tu ticket para {entrada.activity.name}.\n¡Nos vemos el {entrada.activity.startDateandtime}!",
        from_email="no-reply@miapp.com",
        to=[data['email']],
    )


    # Adjuntar PDF
    for i, ticket in enumerate(tickets):
    # ticket es un BytesIO o archivo
        ticket.seek(0)
        email.attach(f"ticket_{i+1}.pdf", ticket.read(), "application/pdf")
  #  email.attach("ticket.pdf", tickets.read(), "application/pdf")

    # Enviar
    email.send()

    return Response(status=201)


@api_view(['GET'])
def user_profile(request):

    from_QR = request.GET.get('fromQR', None)
    if not request.user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return Response({'error': 'Profile not found'}, status=404)
    

    if not from_QR:
        serializer = UserProfileSerializer(user_profile, context={'request': request})
    else:
        serializer = UserProfileSerializerForQR(user_profile, context={'request': request})

    return Response(serializer.data)

@api_view(['POST'])
def update_profile_image (request):

    if not request.user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)
    
    try:
        userProfile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return Response({'error': 'Profile not found'}, status=404)
    serializer = UserProfileSerializer(userProfile, data=request.data, partial=True)

    if serializer.is_valid():

        serializer.save()
        return Response(serializer.data, status=200)
    return Response(serializer.errors, status=400)


@api_view(['POST'])
def remove_profile_image (request):

    if not request.user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)
    
    try:
        userProfile = UserProfile.objects.get(user=request.user)
        if userProfile.image:
            userProfile.image.delete(save = True)
            userProfile.save()
    except UserProfile.DoesNotExist:
        return Response({'error': 'Profile not found'}, status=404)

    return Response(status=200)
   


@api_view(['POST'])
def updatePassword(request):
    
    user = request.user
    if not user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)
    
    data = request.data.copy()
    
    if not user.check_password(data['current_pwd']):
        return Response({"error": "La contraseña actual es incorrecta."}, status=400)

    if data['new_pwd'] != data['confirm_pwd']:
        return Response({"error": "Las contraseñas no coinciden."}, status=400)

    user.set_password(data['new_pwd'])
    user.save()

    return Response({"message": "Contraseña actualizada correctamente."}, status=200)

import calendar



from itertools import chain, groupby


@api_view(['GET'])
def billingStatus(request):
    user = request.user
    if not user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)
    
    try:
        userProfile = UserProfile.objects.get(user = user)
    except UserProfile.DoesNotExist:
        return Response(400)
    
    first_date, last_date = get_month_bounds()

    activities = Activity.objects.filter(creador__user = user, startDateandtime__date__range=(first_date, last_date))
    promos = Promo.objects.filter(creador__user=user, startDateandtime__date__range=(first_date,last_date))

    #paymentReceipt = PaymentForUse.objects.filter(userProfile__user = user, payment_month__date__range = (first_date, last_date))
    
    
    #print('pyament receipt')
    #if paymentReceipt.exists():
    #    paymentStatus = 'Pagado'
    #else:
    #    paymentStatus = 'Pendiente de pago'
    #eventos = list(chain(activities, promos))
    
    eventos_gratuitos = 0
    eventos_pagos = 0
    #eventos_sin_centralizar = []
    
    eventos_gratuitos = eventos_gratuitos + promos.count() #S'inicialitza amb la quantitat de promos, que sempre són gratis
    for evento in activities:
        if evento.gratis == True and evento.active== True: #Eventos gratuitos
          
            eventos_gratuitos += 1
        elif evento.active== True: #Eventos de pago:
            eventos_pagos += 1
       # elif evento.control_entradas == True:
       #     eventos_centralizados += 1
       # else:
       #     eventos_sin_centralizar.append(evento)

    # Mostrar el tramo de los planes pagos centralizados

    
        #0 a 12 --> 20.000ARS
        #13 a 18 --> 40.000ARS
        #19 a 24 --> 60.000ARS
        #+25 --> 75.000ARS
    centralizado_start_range = userProfile.payment_events_range.start_range if userProfile.payment_events_range else 1
    centralizado_end_range = userProfile.payment_events_range.end_range if userProfile.payment_events_range else 4
    price_range_centralizados = userProfile.payment_events_range.price   if userProfile.payment_events_range else None 

    #ranges = PaymentEventsRanges.objects.all().order_by('start_range')
    #print('ranges: ', ranges)

    #for range in ranges:
    #    if eventos_pagos >= range.start_range and (eventos_pagos <= range.end_range or range.end_range is None):
    #        centralizado_start_range = range.start_range
    #        centralizado_end_range = range.end_range
    #        price_range_centralizados = range.price
    #if eventos_pagos < 5 :
    #    centralizado_end_range = 4
    #    price_range_centralizados = 20000
    #elif eventos_pagos < 9:
    #    centralizado_start_range = 5
    #    centralizado_end_range = 8
    #    price_range_centralizados = 35000
    #elif eventos_pagos < 13:
    #    centralizado_start_range = 9
    #    centralizado_end_range = 12
    #    price_range_centralizados = 50000
    #elif eventos_pagos < 19:
    #    centralizado_start_range = 13
    #    centralizado_end_range = 18
    #    price_range_centralizados = 65000
    #else:
    #    centralizado_start_range = 17
    #    centralizado_end_range = None
    #    price_range_centralizados = 100000

    # Agrupamos sin centralizar por rangos de precios
   # price_ranges = {
   #     "0-24999": 0,
   #     "25-49999": 0,
   #     "50000-99999": 0,
   #     "100+":0
   # }
   # price_external_eventos = 0

   # price_range_costos = {
   #     "0-24999": 0,
   #     "25-49999": 0,
   #     "50000-99999": 0,
   #     "100+":0
   # }


    #price_ranges_costos_totales = {
    #    "0-24999": 15000,
    #    "25-49999": 20000,
    #    "50000-99999": 35000,
    #    "100+":50000
    #}

    #for external_evento in eventos_sin_centralizar:
    #    if external_evento.price < 25000:
    #        price_ranges['0-24999'] += 1
    #        price_external_eventos += 15000
    #        price_range_costos['0-24999']+= 15000
    #    elif external_evento.price < 50000:
    #        price_ranges["25-49999"] += 1
    #        price_external_eventos += 20000
    #        price_range_costos["25-49999"]+= 20000
    #    elif external_evento.price < 100000:
    #        price_ranges["50000-99999"] +=1
    #        price_external_eventos += 35000
    #        price_range_costos['50000-99999']+= 35000
    #    else:
    #        price_ranges["100+"] +=1
    #        price_external_eventos += 50000
    #        price_range_costos['100+']+= 50000


    
    total_goLocal_mensual = price_range_centralizados
    

    #booleano para mostrar o no botón para planificación del mes siguiente
    show_button = False
    fecha_hoy = date.today()
    ultimo_dia_este_mes = calendar.monthrange(fecha_hoy.year, fecha_hoy.month)[1]
    if (fecha_hoy.day > ultimo_dia_este_mes - 12) and userProfile.pago_suscripcion_mes_proximo == False:
        show_button = True
        

    return Response({
        "mostrar_boton_planificacion": show_button,
        "gratuitos": eventos_gratuitos,
        "gratuitos_disponibles": userProfile.available_planes_gratis,
        "centralizados": eventos_pagos,
        "range_centralizados":{"start_range":centralizado_start_range, "end_range":centralizado_end_range},
        "price_range_centralizados" : price_range_centralizados,
       # "sin_centralizar": {
       #     "total": len(eventos_sin_centralizar),
       #     "por_rango": price_ranges
       # },
      #  "price_external_eventos": price_external_eventos,
      #  "sin_centralizar_costos":price_ranges_costos_totales,
      #  "sin_centralizar_costos_por_rango" : price_range_costos,
        "precio_total":total_goLocal_mensual,
        "estado":'Pagado'

    }, status = 200)


def get_month_bounds():
    today = date.today()
    # Primer día del mes
    first_day = today.replace(day=1)
    
    # Último día del mes
    last_day_num = calendar.monthrange(today.year, today.month)[1]
    last_day = today.replace(day=last_day_num)
    
    return first_day, last_day


@api_view(['POST'])
def ableToTurnVisible(request):

    user = request.user
    if not user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)

    try:
        userProfile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response('Such user does not exist', status = 400)
    
    first_date, last_date = get_month_bounds()
    print('first_date: ', first_date)
    print('free_event: ', request.data['free_event'])
    status = None
    if request.data['free_event'] == True: #Planes gratuitos
        print('free event')
        available_free_plans_fromUP = userProfile.available_planes_gratis # valor máximo que tiene para el mes
        activities = Activity.objects.filter(creador__user = user, startDateandtime__date__range=(first_date, last_date), gratis = True, active = True)
        promos = Promo.objects.filter(creador__user=user, startDateandtime__date__range=(first_date,last_date), active = True)
        created_free_events = activities.count() + promos.count()
        available_free_plans = available_free_plans_fromUP-created_free_events
        status = available_free_plans > 0
    else: #Planes pagos
        print('payment event')
        activities = Activity.objects.filter(creador__user=user, startDateandtime__date__range=(first_date, last_date), gratis = False, active = True)
        status = activities.count() < userProfile.payment_events_range.end_range
    print('status')

    return Response(status, status = 200)   


@api_view(['GET'])
def bonos(request):
  
    try:
        bonos = Bono.objects.all().order_by('price')
    except Bono.DoesNotExist:
        return Response('No available bonos at the moment', status = 400)

    serializer = BonoSerializer(bonos, many = True)

    return Response(serializer.data, status = 200)

    
@api_view(['GET'])
def paymentEventsRanges(request):
  
    try:
        paymentEventsRanges = PaymentEventsRanges.objects.all()
    except PaymentEventsRanges.DoesNotExist:
        return Response('No available payment events ranges at the moment', status = 400)

  
    serializer = PaymentEventsRangesSerializer(paymentEventsRanges, many = True)

    return Response(serializer.data, status = 200)



@api_view(['POST'])
def user_prorfile_from_uuid(request):
    user_uuid = request.data['user_uuid']
    try:
        user = UserProfile.objects.get(uuid=user_uuid)
    except UserProfile.DoesNotExist:
        return Response('Not such userprofile', status = 404)

    serializer = UserProfileSerializer(user, context={'request': request})
    return Response(serializer.data) 

@api_view(['GET'])
def get_tickets(request):
    
    user = request.user

    if not user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)

    activity_uuid = request.GET.get('activity_uuid')
    try:
        activity = Activity.objects.get(uuid = activity_uuid)
    except Activity.DoesNotExist:
        return Response('Such activity does not exist', status = 404)

    try: 
        tickets = Ticket.objects.filter(
            entrada__activity = activity,
            user_profile__user = user
        )
    except:
        return Response('You have no tickets for such activity ${activity_uuid}', status=404)
    serializer = TicketSerializer(tickets, many= True, context={'request': request}, fields = ['entrada', 'nombre', 'email', 'qr_code'])
    return Response(serializer.data)

@api_view(['GET'])
def transactions(request):

    user = request.user

    if not user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=404)
    
    try:
        userProfile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response({'error':'Error retrieving UserProfile'}, status = 400)
    

    incoming_payments = userProfile.incoming_payments.all()
    incoming_payments_list =  list(incoming_payments.values('activity', 'status', 'total_amount','payout_status','descripcion','updated_at','payment_id')) 
    approved_payments = list(filter(lambda x: x['status'] == 'approved', incoming_payments_list)) #pagos realizados por los usuarios
    approved_payments = list(filter(lambda x: x['activity']!= None, approved_payments)) #pagos destinados a pagar entradas, no internos
    app_service_payments = list(filter(lambda x:x['activity'] == None, incoming_payments_list))
    total_amount_app_service = sum(payment['total_amount'] for payment in app_service_payments)

    approved_payments.sort(key=lambda x: x['activity']) 

    total_amount_pending = 0
    total_amount_approved = 0

    # Iteramos sobre la lista y sumamos según el status
    for payment in approved_payments:
        if payment['payout_status'] == 'pending':
            total_amount_pending += payment['total_amount']
        elif payment['payout_status'] == 'approved':
            total_amount_approved += payment['total_amount']
    
    total_amount_global = total_amount_pending + total_amount_approved
    


    grouped_by_activity = groupby(approved_payments, key=lambda x: x['activity']) #agrupar por actividad

    payments_grouped_dict = {}

    for activity, group in grouped_by_activity:
        payments_grouped_dict[activity] = list(group)

    data_grouped_dict = {}
    next_due_date = None
    due_date_amount = 0
    for activity in payments_grouped_dict: #montar estructura de datos
        total_amount = 0
        activity_obj= Activity.objects.get(id=activity)
        for payment in payments_grouped_dict[activity]:
            total_amount += payment['total_amount']
        
        if  payments_grouped_dict[activity][0]['payout_status'] != 'approved' and next_due_date == None:
            print('entra')
            next_due_date = activity_obj.startDateandtime + timedelta(days=1) 
            due_date_amount = total_amount
    
        elif payments_grouped_dict[activity][0]['payout_status'] != 'approved' and (activity_obj.startDateandtime + timedelta(days=1)).date() <= next_due_date.date(): 
            print('entra a la suma: ', payments_grouped_dict[activity][0])
            next_due_date = activity_obj.startDateandtime + timedelta(days=1) 
            due_date_amount += total_amount
            
        data_grouped_dict[activity] = {"total_amount": total_amount, "due_date": activity_obj.startDateandtime + timedelta(days=1), "status": payment['payout_status'],"activity": {"uuid": activity_obj.uuid, "image": activity_obj.image.url if activity_obj.image else None, "name":activity_obj.name,"startDateandtime":activity_obj.startDateandtime}}
     
    grouped_by_status = defaultdict(list)
    for activity, data in data_grouped_dict.items():
        status = data['status']  # Obtén el status
        grouped_by_status[status].append(data)
    grouped_by_status = dict(grouped_by_status)

    return Response({'payments':grouped_by_status, 'app_service_payments':app_service_payments,'total_amounts':{'total_amount': total_amount_global, 'payed_amount': total_amount_approved, 'pending_amount': total_amount_pending, 'app_service_payments': total_amount_app_service, 'next_due_date':next_due_date, 'due_date_amount':due_date_amount}},status = 200)

from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes

from django.contrib.auth.tokens import PasswordResetTokenGenerator
token_generator = PasswordResetTokenGenerator()

@api_view(['POST'])
def sendMailForPwdRecovery(request):
    
    email = request.data.get("email", "").strip().lower()
    if not email:
        return Response({"detail": "Si el email existe, te enviaremos instrucciones."}, status = 200)

    try:
        user = User.objects.get(email = email)

    except User.DoesNotExist:
        #Responta genèrica, sempre la mateixa, per no desvelar si existeix l'usuari o no
        return Response({"detail": "Si el email existe, te enviaremos instrucciones."}, status = 200)
           
    uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
    token = token_generator.make_token(user)

    reset_link = f"https://go-local-web.vercel.app/reset-password?uid={uidb64}&token={token}"

    subject = "Recupera tu contraseña"
    message = f"Hola,\n\nHaz clic en el siguiente enlace para restablecer tu contraseña:\n{reset_link}\n\nSi no fuiste tú, ignora este correo."
    send_mail(subject, message, 'no-reply@miapp.com', [email], fail_silently=False)

    return Response({"detail": "Si el email existe, te enviaremos instrucciones."}, status=200)

    
@api_view(['POST'])
def update_user(request):
    data = request.data.copy()
    user_uuid = data['user_uuid']
    try:
        user = UserProfile.objects.get(uuid=user_uuid)
        if 'description' in data:
            user.bio = data['description']
        if 'bio' in data:
            user.bio = data['bio']
        
        if 'location' in data:
            user.location = data['location']

        if 'tags' in data:
        # Parsear la lista de tags correctamente
            tags_raw = request.data.get('tags', '[]')
            try:
                tags_list = ast.literal_eval(tags_raw)
                if not isinstance(tags_list, list):
                    tags_list = []
            except (ValueError, SyntaxError):
                tags_list = []
            user.tags.set(tags_list)
        
        if 'image' in data:
            user.image = data['image']

        if 'telefono' in data:       
            user.telefono = int(data['telefono'])
        
        if 'email' in data:
            user.user.email = data['email']
            user.user.save()
        if 'asCreator' in data:
            user.creador = bool(data['asCreator'])

        user.save()
        return Response({'user_uuid': user.uuid}, status=200)

    except UserProfile.DoesNotExist:
        return Response('No such UserProfile exists', status=404)
    
def find_nearest_place(lat, lng):
    all_places = Place.objects.all()
    for place in all_places:
        dist = geodesic((lat, lng), (place.latitude, place.longitude)).km
        if dist < 50: 
            return place

    return None

@api_view(['POST'])
def sign_in(request):

    data = request.data

    if User.objects.filter(email=data['email']).exists():
        return Response({"email_error":"Este mail ya está en uso para un usuario"},status=404)
    if User.objects.filter(username=data['name']).exists():
        return Response({"name_error":"Este nombre ya está en uso para un usuario"},status=404)
    
    user = User.objects.create_user(
            username=data['name'],
            email=data['email'],
            password=data['password'],
            first_name=data.get('name', '')
        )
    userProfile = UserProfile.objects.create(user=user)

    refresh = RefreshToken.for_user(user)
    return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'user_uuid':str(userProfile.uuid)
        })

@api_view(['POST'])
def google_sign_in(request):
    id_token = request.data.get('idToken')
    if not id_token:
        return Response({"error": "Token requerido"}, status=400)
    response = request.get(f"https://oauth2.googleapis.com/tokeninfo?id_token={id_token}")
    if response.status_code != 200:
        return Response({"error": "Token inválido"}, status=400)

    payload = response.json()
    email = payload['email']
    name = payload.get('name', '')

    user, created = User.objects.get_or_create(
        email=email,
        defaults={'username': email, 'first_name': name}
    )
    userProfile = UserProfile.objects.create(user=user)
    refresh = RefreshToken.for_user(user)
    return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
        })


@api_view(['POST'])
def updateActivityAssistance(request):

    data = request.data

    user = request.user
    activity_uuid = data.get('activity_uuid')
    promo_uuid = data.get('promo_uuid')


    if activity_uuid:
        try:
            activity = Activity.objects.get(uuid=activity_uuid)
            try:
                update_value = int(data['update'])
                try:
                        userProfile = UserProfile.objects.get(user=user)
                        if userProfile:
                            if update_value == 1 : 
                                userProfile.activities.add(activity)
                            elif update_value == -1:
                                userProfile.activities.remove(activity)

                except UserProfile.DoesNotExist:
                    return Response('Error in access token', status = 404)
                
            except (ValueError, TypeError):
                return Response({'error': 'update must be an integer'}, status=400)
            
            asistentes = UserProfile.objects.filter(activities=activity)
            serializer = UserProfileBasicSerializer(asistentes, many=True, context={"request": request})

            return Response(serializer.data, status=200)


        except Activity.DoesNotExist:
            return Response({'error': 'Activity not found'}, status=404)
    elif promo_uuid:
        try:
            promo = Promo.objects.get(uuid=promo_uuid)
            try:
                update_value = int(data['update'])
                try:
                    userProfile = UserProfile.objects.get(user=user)
                    if userProfile:
                        if update_value == 1 : 
                            userProfile.promos.add(promo)
                        elif update_value == -1:
                            userProfile.promos.remove(promo)

                except UserProfile.DoesNotExist:
                    return Response('Error in access token', status = 404)
            except (ValueError, TypeError):
                return Response({'error': 'update must be an integer'}, status=400)

            promo.asistentes = promo.asistentes + update_value
            promo.save()

            return Response({'asistentes': promo.asistentes}, status=200)


        except Promo.DoesNotExist:
            return Response({'error': 'Activity not found'}, status=404)


@api_view(['POST'])
def create_event(request):
    user = request.user
   
    if not user.is_authenticated:
        return Response({'error': 'User is not authenticated'}, status=401)

    try: 
        userProfile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response('Unable to retrieve UserProfile from User instance', status = 403)

    data = request.data
    if isinstance(data, QueryDict):
        data._mutable = True
    lat = float(request.data['lat'])
    lng = float(request.data['lng'])
    nearest_place = find_nearest_place(lat, lng)
    is_private_plan = request.data['tipoEvento'] == '2'
    tipoEvento = int(request.data['tipoEvento'])
    tags_list = request.data['tags']
    if is_private_plan == False:
        tags_raw = request.data['tags']
        
        try:
            tags_list = ast.literal_eval(tags_raw)
            if not isinstance(tags_list, list):
                    tags_list = []
        except (ValueError, SyntaxError):
            tags_list = []
            
        del data['tags'] #ho borrem del serializer d'Activity i els passarem després
    
    data['lat'] = float(lat)
    data['long'] = float(lng)
        
    if not nearest_place:
            ##send_mail(
            ##    'Nueva actividad sin lugar asociado!',
            ##   f'Se creó una actividad sin place. Coordenadas: {lat}, {lng}. El nombre de la actividad es:{name}',
            ##    'rebuig.lauragonzalez@gmail.com',
            ##    ['lauragonzalezcomalada@gmail.com'],
            ##    fail_silently=False,
            ##)
        data['place'] = None
    else:
        data['place'] = nearest_place.id

    data['gratis'] = request.data.get('gratis') == 'true' #if request.data.get('gratis') == 'true' == 'True': True; si es 'false' == 'true' --> False
    
    if data['gratis'] == True:
        data['reserva_necesaria'] = request.data.get('reserva') == 'true'

    if (data['gratis'] == False and tipoEvento == 0): #gratis = False i es activity
        data['control_entradas'] = request.data.get('centralizarEntradas') == 'true'
        
        if(data['control_entradas'] == False):
            data['tickets_link'] = request.data.get('tickets_link', None)
    
    data.pop('tipoEvento', None)
    if tipoEvento == 0:
        serializer = ActivitySerializer(data=data)
    elif tipoEvento == 1:
        serializer = PromoSerializer(data=data)
    else: #data['tipoEvento] == '2'
        serializer = PrivatePlanSerializer(data=data)

    if serializer.is_valid():
        event = serializer.save()
        if is_private_plan == False:
        
            if tags_list:
                event.tags.set(tags_list)

            if  data['gratis'] ==  True:
                if(data['reserva_necesaria'] == True):    
                    reservas_list = json.loads(data['reservas'])
                    for reserva in reservas_list:
                        tipo = reserva.get('tipoReserva', '')
                        cantidad = int(reserva.get('cantidad', 0))
                        campos_data = reserva.get('campos', [])

                        # Extraemos los UUIDs y buscamos instancias existentes de CampoReserva
                        campo_uuids = [campo.get('uuid') for campo in campos_data]
                        campos_queryset = CampoReserva.objects.filter(uuid__in=campo_uuids)

                        if tipoEvento == 0:
                            reserva_instance = ReservaForm.objects.create(
                                nombre=tipo,
                                activity=event,
                                max_disponibilidad=cantidad
                            )
                        elif tipoEvento == 1:
                              reserva_instance = ReservaForm.objects.create(
                                nombre=tipo,
                                promo=event,
                                max_disponibilidad=cantidad
                            )
                        reserva_instance.campos.set(campos_queryset)
                        reserva_instance.save()

                        # Asociamos la reserva con el evento
                        event.reservas_forms.add(reserva_instance)
            elif data['gratis'] == False and data['control_entradas'] == True and tipoEvento == 0:
               
                entradas_list = json.loads(data['entradas'])
                for entrada in entradas_list:
                    entradaCreada = EntradasForPlan.objects.create(
                        titulo=entrada.get('nombre'),
                        desc=entrada.get('descripcion'),
                        precio=float(entrada.get('precio')),
                        maxima_disponibilidad=float(entrada.get('cantidad'))
                    )
                    entradaCreada.save()
                    event.entradas_for_plan.add(entradaCreada)

                    
                #[{"nombre":"ENTRADA GALLINERO1","descripcion":"Sin vision","precio":"10000","cantidad":"10"},
                # {"nombre":"Entrada VIP","descripcion":"Conoce al artista","precio":"30000","cantidad":"5"}]


        event.creador = userProfile
        
        event.save()
        return Response({'uuid':event.uuid, 'tipo': tipoEvento}, status=201)
    else:
        print('serializer errors:', serializer.errors)
    return Response(serializer.errors, status=400)




@api_view(['GET'])
def search_users(request):

    query = request.GET.get('query', '')
    results = UserProfile.objects.filter(user__username__icontains=query)[:10]

    serializer = UserProfileSerializer(results, many=True, fields=['uuid','username'] )
    return Response(serializer.data, status = 200)


@api_view(['GET'])
def private_plans(request):

    user = request.user
    privatePlanUuid = request.GET.get('privatePlanUuid','')
    print('privatePlanUuid: ', privatePlanUuid)
    #Generic per tenir els privatePlans d'un user
    if not privatePlanUuid:
        try:
            userProfile = UserProfile.objects.get(user = user)
        except UserProfile.DoesNotExist:
            return Response(status = 404)   
        
        serializer = PrivatePlanSerializer(userProfile.planes_invitados.all(), many=True,context={'request': request})
    # return Response(UserProfileSerializer(userProfile).data, status = 200)
        return Response(serializer.data, status = 200)
    
    #Per tenir el detall d'un privatePlan
    else:
        try:
            privatePlan = PrivatePlan.objects.get(uuid=privatePlanUuid)
        except PrivatePlan.DoesNotExist:
            return Response('No hay tal plan para tal usuario',status = 404)

        serializer = PrivatePlanSerializer(privatePlan, context={'request': request})
        return Response(serializer.data, status = 200)
    
@api_view(['POST'])
def update_item_details(request):
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    try:
        userProfile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response('User profile not found', status=404)
    
    data = request.data.copy()
    item_uuid = data['uuid']
    private_plan_uuid = data['plan_uuid']
    amount = data['newAssignedAmountValue']

    try:
        item = ItemPlan.objects.get(uuid=item_uuid, private_plan__uuid=private_plan_uuid, private_plan__invited_users__user=user)
    
    except ItemPlan.DoesNotExist:
        return Response('No hay tal plan')
    

    item.assignedAmount = float(amount)
    item.people_in_charge.add(userProfile)
    item.save()

    try: 
        privatePlan = PrivatePlan.objects.get(uuid=private_plan_uuid, invited_users__user=request.user)
    except PrivatePlan.DoesNotExist:
        return Response('No hay tal plan, o no estas invitado...', status=404)

    return Response(PrivatePlanSerializer(privatePlan).data, status = 200)

@api_view(['POST'])
def add_item(request):

    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    data = request.data.copy()
    private_plan_uuid = data['plan_uuid']

    try:
        privatePlan = PrivatePlan.objects.get(uuid=private_plan_uuid,
                                                invited_users__user=user)

    except PrivatePlan.DoesNotExist:
        return Response('No hay tal plan, o no estas invitado...', status=404)
    
    
    item = ItemPlan.objects.create(
        private_plan=privatePlan,
        name=data['name'],
        neededAmount=float(data['neededAmount']))
    


    return Response(ItemSerializer(item).data, status = 200)

@api_view(['POST'])
def accept_invitation(request):
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    private_plan_uuid = request.data.get('private_plan_uuid')
    if not private_plan_uuid:
        return Response('Private plan UUID is required', status=400)
    
    try:
        private_plan = PrivatePlan.objects.get(uuid=private_plan_uuid)
    except PrivatePlan.DoesNotExist:
        return Response('Private plan not found', status=404)   
    
    private_plan_invitation = PrivatePlanInvitation.objects.create(
        private_plan=private_plan,
        invited_user=user.profile,
        status=1  # Aceptado
    )
    return Response(status=200)


def invitation_redirect(request, invitation_code):

    try:
        privatePlan = PrivatePlan.objects.get(invitation_code=invitation_code)
    except PrivatePlan.DoesNotExist:
        return render(request, 'error.html', {'message': 'Plan not found or invitation code is invalid.'})

    # Redirecciona al deep link si es móvil
    #deep_link = f"goLocal://privatePlan/{privatePlan.uuid}"
    deep_link = f"golocal://privateplaninvitation/{privatePlan.uuid}"
    web_fallback_url = f"http://localhost:8000/plan/{invitation_code}"

    user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
    is_mobile = 'android' in user_agent or 'iphone' in user_agent

    # Si no es móvil, mostrás una landing web o redireccionás al sitio
    # S'ha de crear aquesta landing web
    return render(request, 'deep_link_landing.html', {
        'deep_link': deep_link,
        'web_fallback_url': web_fallback_url,
        'is_mobile': is_mobile,
    })

#obtenir la setmana en curs
#def get_week_range():
#    today = datetime.now().date()
#    start_of_week = today - timedelta(days=today.weekday())  # lunes
#    end_of_week = start_of_week + timedelta(days=6)  # domingo
#    return start_of_week, end_of_week

## MÉTODO PARA goLocalWeb
## Obtenir els events creats x user per la setmana en curs


def get_week_range_by_date(date_obj):
    # Ajusta al lunes de esa semana (considerando lunes como inicio)
    weekday = date_obj.weekday()  # 0=lunes, 6=domingo
    monday = date_obj - timedelta(days=weekday)
    sunday = monday + timedelta(days=6)
    return monday.date(), sunday.date()


@api_view(['GET'])
def userCreatedEventsForTheWeek(request):

    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    # Leer parámetro date YYYY-MM-DD desde query params
    date_str = request.query_params.get('date')  # ejemplo: "2025-08-11"
    if date_str:
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        except ValueError:
            return Response('Invalid date format, use YYYY-MM-DD', status=400)
    else:
        # si no viene, usamos hoy
        date_obj = datetime.today()

    start_of_week, end_of_week = get_week_range_by_date(date_obj)

    activities = Activity.objects.filter(creador__user=user, startDateandtime__date__range=(start_of_week, end_of_week))
    promos = Promo.objects.filter(creador__user=user, startDateandtime__date__range=(start_of_week, end_of_week))
    private_plans = PrivatePlan.objects.filter(creador__user=user,startDateandtime__date__range=(start_of_week, end_of_week))

    data = {
        'activities': ActivitySerializer(activities, many=True, fields=['uuid','name','shortDesc','place_name','image','startDateandtime','tag_detail','gratis','creador_image','asistentes','active']).data,
        'promos': PromoSerializer(promos, many=True, fields=['uuid','name','shortDesc','place_name','image','startDateandtime','tag_detail','gratis','creador_image','asistentes','active']).data,
        'private_plans': PrivatePlanSerializer(private_plans, many=True, fields=['uuid','name','shortDesc','image','startDateandtime','creador_image','asistentes','active']).data,
    }
    return Response(data)

@api_view(['GET'])
def entradasForUserAdmin(request):
    user = request.user
    resultados = []

    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    try:
        userProfile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response('Error in obtaining UserProfile', status=400)
    def procesar_evento(evento, tipo):
        entradas = EntradasForPlan.objects.filter(**{evento_field_map[tipo]: evento})
        entradas_data = []
        entradas_con_tickets = entradas.annotate(ticket_count=Count('tickets'))

        for entrada in entradas_con_tickets:
            total = entrada.maxima_disponibilidad or 0
            vendidos = entrada.maxima_disponibilidad - entrada.disponibles if total else 0
            porcentaje = (vendidos / total * 100) if total else 0

            dinero_recaudado = entrada.tickets.count() * entrada.precio  # ya que Ticket no tiene `price`, se usa entrada.precio
            vendidas = entrada.tickets.count()

            entradas_data.append({
                'entrada_uuid':entrada.uuid,
                'entrada_name': entrada.titulo,
                'entrada_shortDesc': entrada.desc,
                'entrada_precio': entrada.precio,
                'entrada_max_disp': entrada.maxima_disponibilidad,
                'entrada_disponibles': entrada.disponibles,
                'entrada_vendidas':vendidas,
                'entrada_porcentajedeventas': round(porcentaje, 2),
                'entrada_dinero_recaudado': dinero_recaudado,
            })

        if entradas_data:
            resultados.append({
                'views': evento.views,
                'shares': evento.shares,
                'tracking_tipo':0,
                'tipo': tipo,
                'event_name': evento.name,
                'event_uuid': evento.uuid,
                'entradas': entradas_data,
                'event_dateandtime': evento.startDateandtime,
                'event_imageUrl': evento.image.url if evento.image else None

            })


    def procesar_reservas(evento, tipo):
        reservas = ReservaForm.objects.filter(**{evento_field_map[tipo]: evento})
        reservas_data = []
       # entradas_con_tickets = entradas.annotate(ticket_count=Count('tickets'))
        for reserva in reservas:
            #hechas = Reserva.objects.filter(reserva_form = reserva)
            reservados = reserva.confirmados
            porcentaje_reservado = (reservados/reserva.max_disponibilidad)*100
         
            reservas_data.append({
                'uuid':reserva.uuid,
                'nombre':reserva.nombre,
                'max_disponibilidad':reserva.max_disponibilidad,
                'confirmadas': reservados,
                'porcentaje_reservado' : round(porcentaje_reservado,2),
                'campos': CampoReservaSerializer(reserva.campos, many = True).data,
            })

        if reservas_data:
            resultados.append({
                'tracking_tipo':1,
                'views': evento.views,
                'shares': evento.shares,
                'tipo': tipo,
                'event_name': evento.name,
                'event_uuid': evento.uuid,
                'reservas': reservas_data,
                'event_dateandtime': evento.startDateandtime,
                'event_imageUrl': evento.image.url if evento.image else None

            })

    evento_field_map = {
        0: 'activity',
        1: 'promo',
        2: 'privateplan'
    }
    # Activities
    for activity in Activity.objects.filter(creador=userProfile, active = True):
        if EntradasForPlan.objects.filter(activity=activity).exists():
            procesar_evento(activity, tipo=0)

        elif ReservaForm.objects.filter(activity=activity).exists():
            procesar_reservas(activity,tipo = 0)

        elif activity.gratis == True and activity.reserva_necesaria == False:
            resultados.append({'tracking_tipo':2,
                               'views': activity.views,
                               'shares': activity.shares,
                               'tipo':0,
                               'event_uuid':activity.uuid,
                               'event_name':activity.name,
                               'event_dateandtime': activity.startDateandtime,
                                'event_imageUrl': activity.image.url if activity.image else None
                               })
        elif activity.gratis == False and activity.control_entradas == False:
            resultados.append({
                'tracking_tipo':3,
                'views': activity.views,
                'shares': activity.shares,
                'clicks_tickets_link':activity.clicks_on_tickets_link,
                'tickets_link': activity.tickets_link,
                'tipo':0,
                'event_uuid': activity.uuid,
                'event_name': activity.name,   
                'event_dateandtime': activity.startDateandtime,
                'event_imageUrl': activity.image.url if activity.image else None
            })
    # Promos
    for promo in Promo.objects.filter(creador=userProfile, active=True):
        if ReservaForm.objects.filter(promo=promo).exists():
            procesar_reservas(promo, tipo=1)
            
        elif promo.reserva_necesaria == False:
            resultados.append({'tracking_tipo':3,
                               'views': promo.views,
                               'shares': promo.shares,
                               'tipo':0,
                               'event_uuid':promo.uuid,
                               'event_name':promo.name,
                               'event_dateandtime': promo.startDateandtime,
                                'event_imageUrl': promo.image.url if promo.image else None})

    # PrivatePlans
    for plan in PrivatePlan.objects.filter(creador=userProfile):
        if EntradasForPlan.objects.filter(privateplan=plan).exists():
            procesar_evento(plan, tipo=2)

    def ordenar_por_fecha(item):
        start = item['event_dateandtime']
        start_date = start.date()
        today = date.today()
        if start_date == today:
            return (0, start)  # primero los de hoy
        elif start_date > today:
            return (1, start)  # luego los futuros
        else:
            return (2, start)
        
    resultados_ordenados = sorted(resultados, key=ordenar_por_fecha)
    # Agrupar por fecha (formato 'dd/mm')
    agrupados_temp = defaultdict(list) 
    for evento in resultados_ordenados:
        fecha_str = evento['event_dateandtime'].strftime('%d/%m') if evento['event_dateandtime'] else 'Sin fecha'
        agrupados_temp[fecha_str].append(evento)

    def fecha_clave(fecha_str):
        try:
            dia, mes = map(int, fecha_str.split('/'))
            fecha = date(date.today().year, mes, dia)
        except:
            return (3, date.max)  # para 'Sin fecha' u otras fallas
        today = date.today()
        if fecha == today:
            return (0, fecha)
        elif fecha > today:
            return (1, fecha)
        else:
            return (2, fecha)

    # Crear OrderedDict con las fechas agrupadas y ordenadas
    agrupados_por_fecha_ordenado = OrderedDict(
        sorted(agrupados_temp.items(), key=lambda item: fecha_clave(item[0]))
    )


    return Response(agrupados_por_fecha_ordenado)

from collections import defaultdict


@api_view(['POST'])
def soldTicketsForEvent(request):
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)

    data = request.data.copy()
    event_uuid = data['event_uuid']
    tipo = data['event_type']
    tipo_tickets = data['tickets_type']
    tickets = []

    if tipo_tickets == 0 and tipo == 0: #means Tickets y Activity. Solo se pueden comprar tickets en las activities
       
        try:
            tickets = Ticket.objects.filter(entrada__activity__uuid = event_uuid, entrada__activity__creador__user = user)
        except Ticket.DoesNotExist:
            return Response('No hi ha tickets per aquesta activitat', status = 404)
        
        serializer = TicketSerializer(tickets, many = True)
        return Response(serializer.data, status = 200)

    elif tipo_tickets == 1: #means Reservas
        if tipo == 0: #means activity amb reserva
            try: 
               reservas = Reserva.objects.filter(
                    reserva_form__activity__uuid=event_uuid
                ).select_related('reserva_form')

            except Reserva.DoesNotExist:
                return Response('No hay reservas asociadas a este evento', status = 404)
        
        elif tipo == 1: #means promo amb reserva
            try:
                reservas = Reserva.objects.filter(reserva_form__promo__uuid = event_uuid).select_related('reserva_form')
            except Reserva.DoesNotExist:
                return Response('No hay reservas asociadas a este evento', status = 404)

        else:
            return Response('Error del tipo de evento', 400)
        
        # Agrupar por nombre del formulario
        grupo_reservas = defaultdict(list)
        for r in reservas:
            grupo_reservas[r.reserva_form.nombre].append({
                "status": r.estado,
                "uuid": r.uuid,
                "values": r.values
            })

        # Formar la respuesta final
        data = {
            "type": tipo_tickets,
            "data": grupo_reservas  # defaultdict se serializa como dict normal
        }
        return Response(data)

@api_view(['POST'])
def updateEvento(request):
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    data = request.data
    tipo = data['event_type']

    try:
        if tipo == 0:
            evento = Activity.objects.get(uuid = data['uuid'])   
            print('tinc evento')
            print(evento)
        elif tipo == 1:
            evento = Promo.objects.get(uuid = data['uuid'])
            print('tinc promo')
    except Activity.DoesNotExist or Promo.DoesNotExist:
        return Response('Error retrieving object', status = 400)
    
    evento.name = data['name']
    evento.shortDesc = data['shortDesc']
    evento.desc = data['desc']
    evento.save()

    if tipo == 0:
        serializer = ActivitySerializer(data = evento)
    elif tipo == 1:
        serializer = PromoSerializer(data = evento)

    return Response(serializer.data, status = 200)

@api_view(['POST'])
def updateEntrada(request):
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    data = request.data.copy()

    entrada_uuid= data['uuid']
    tipo = data['event_type']
    
    entrada = []
    try:
        if tipo == 0:
            entrada = EntradasForPlan.objects.get(uuid = entrada_uuid, activity__creador__user = user)
      #  elif tipo == 1: 
      #      entrada =  EntradasForPlan.objects.get(uuid=entrada_uuid, promo__creador = user)
      #  elif tipo == 2:
      #      entrada = EntradasForPlan.objects.get(uuid = entrada_uuid, privatePlan__user = user)
    except EntradasForPlan.DoesNotExist:
        return Response('No existeix el tipus de entrada', status = 404)
    
    entrada.titulo = data['titulo']
    entrada.desc = data['descripcion'] 
    entrada.precio = data['precio']
    diferencia_disponibilidades = entrada.disponibles - int(data['disponibles'])
    entrada.disponibles = int(data['disponibles'])
    # si eren 6 i ara en son 8 max_disp = max_disp + 2
    # si eren 6 i ara en son 4, max_disp = max_disp - 2
    entrada.maxima_disponibilidad =  entrada.maxima_disponibilidad-diferencia_disponibilidades
    entrada.save()

    porcentaje = (entrada.maxima_disponibilidad - entrada.disponibles) / entrada.maxima_disponibilidad * 100

    return Response({ "max_disponibilidad": entrada.maxima_disponibilidad, "porcentaje_ventas": round(porcentaje, 2)}, status = 200)


@api_view(['POST'])
def updateReserva(request):
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
        
    data = request.data.copy()
    reserva_uuid= data['uuid']
    tipo = data['event_type']

    try:
        if tipo == 0:
            reserva = ReservaForm.objects.get(uuid = reserva_uuid, activity__creador__user = user)
        elif tipo == 1: 
            reserva =  ReservaForm.objects.get(uuid=reserva_uuid, promo__creador__user = user)
        
    except EntradasForPlan.DoesNotExist:
        return Response('No existeix el tipus de entrada', status = 404)
    
    
    reserva.nombre = data['nombre']
    reserva.max_disponibilidad = int(data['max_disponibilidad'])
    
    campos_data = data['campos']
    uuids = [c["uuid"] for c in campos_data]

    try:
        campos = CampoReserva.objects.filter(uuid__in = uuids)
    except CampoReserva.DoesNotExist:
        return Response('No se encontraron los campos de reserva', status = 404)
    

    reserva.campos.set(campos)
    reserva.save()
    porcentaje = (reserva.confirmados / reserva.max_disponibilidad )* 100
    return Response({ "max_disponibilidad": reserva.max_disponibilidad, "porcentaje_reservados": round(porcentaje, 2)} ,status = 200)
  

@api_view(['POST'])
def updateTicketsLink(request):

    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    data = request.data.copy()
    
    tipo = data['event_type']
    if tipo == 0:
        try:
            activity = Activity.objects.get(uuid = data['uuid'])
        except Activity.DoesNotExist:
            return Response('No se puede actualizar el link porque no se encuentra la Activity', status = 400)
    else:
        return Response('Error en el tipo del evento que se intenta asociar el link')
    activity.tickets_link = data['tickets_link']
    activity.save()

    return Response({'tickets_link': activity.tickets_link}, status = 200)


@api_view(['POST'])
def obtener_campos_reserva(request):
    data = request.data.copy()
    campos = []
    if data['tipo'] == 0: #activity
        campos = CampoReserva.objects.filter(activity__uuid=data['uuid'], activity__gratis = True, activity__reserva_necesaria = True)
    elif data['tipo'] == 1:#promo
        campos = CampoReserva.objects.filter(promo__uuid=data['uuid'], promo__reserva_necesaria = True)
    else:
        return Response('Tipo de evento no válido', status=400)
    
    serializer = CampoReservaSerializer(campos, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def crear_reserva(request):
    serializer = ReservaSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)

@api_view(['GET'])
def camposReserva(request):
    campos = CampoReserva.objects.all()
    if not campos:
        return Response({'error': 'No reservation fields found'}, status=404)

    serializer = CampoReservaSerializer(campos, many=True)
    return Response(serializer.data)

@api_view(['GET', 'POST'])
def templates(request):
    if request.method == 'GET':
        user = request.user
        if not user.is_authenticated:
            return Response('User not authenticated', status=401)
        
        template_uuid = request.query_params.get('uuid',None)
        if not template_uuid:
            try: 
                templates = EventTemplate.objects.filter(creador = user)
            except EventTemplate.DoesNotExist:
                return Response('There are no templates created by this user', status = 200)
            
            serializer = EventTemplateSerializer(templates, many=True)
        
        elif template_uuid:
            try:
                template_data = EventTemplate.objects.get(uuid = template_uuid, creador = user)

            except EventTemplate.DoesNotExist:
                return Response('Error retrieving data from such template', status = 400)
            
            serializer = EventTemplateSerializer(template_data)
        
        return Response(serializer.data, status = 200)
    
    elif request.method == 'POST':
        user = request.user
        if not user.is_authenticated:
            return Response('User not authenticated', status=401)
        data = request.data.copy()
        template = EventTemplate.objects.create(
            name = data['name'],
            tipoEvento = data['tipoEvento'],
            values = json.loads(data['values']),
            creador = user 
        )
        template.save()

        return Response(status = 200)

@api_view(['POST'])
def updateTicketStatus(request):
    user = request.user

    if not user.is_authenticated:
                return Response('User not authenticated', status=401)
    
    try: 
        ticket = Ticket.objects.get(uuid = request.data['ticket_uuid'])
    except Ticket.DoesNotExist:
        return Response('Hubo un error localizando tu ticket', status = 400)

    ticket.status = 1 if ticket.status == 0 else 0
    ticket.save()

    return Response(ticket.status, status=200)

    


@api_view(['POST'])
def updateReservaStatus(request):
    user = request.user

    if not user.is_authenticated:
                return Response('User not authenticated', status=401)
    
    data = request.data.copy()

    try: 
        reserva = Reserva.objects.get(uuid = data['reserva_uuid'])
    except Reserva.DoesNotExist:
        return Response('Such Reserva does not exist', status = 400)

    reserva.estado = int(data['reserva_status'])
    reserva.save()

    return Response(reserva.estado, status = 200)

@api_view(['POST'])
def updateActiveStatus(request):
    user = request.user

    if not user.is_authenticated:
        return Response('User not authenticated', status=401)
    
    try:
        userProfile = UserProfile.objects.get(user = user)
    except UserProfile.DoesNotExist:
        return Response('User profile not found', status = 404)
    

    data = request.data.copy()
    if data['tipo'] == 0: #activity
        try:
            evento = Activity.objects.get(uuid = data['uuid'])
        except Activity.DoesNotExist:
            return Response('Actividad no encontrada', status = 400)
        
    elif data['tipo'] == 1: #promo
        try:
            evento = Promo.objects.get(uuid = data['uuid'])
        except Promo.DoesNotExist:
            return Response('Promo no encontrada', status = 400)
        
    else: #private plan
        try: 
            evento = PrivatePlan.objects.get(uuid = data['uuid'])
        except PrivatePlan.DoesNotExist:
            return Response('Private plan no encontrado', status = 400)


    active_status_previo = evento.active
    evento.active = not active_status_previo
    evento.save()

    if evento.gratis == True:
        if userProfile.available_planes_gratis <= 0:
            return Response('No tienes más planes gratis disponibles para activar este evento', status = 400)
        userProfile.available_planes_gratis -= 1
        userProfile.save()
    return Response(evento.active, status = 200)
        
    


import openpyxl
from django.http import HttpResponse, QueryDict, QueryDict

@api_view(['GET'])
def export_to_excel(request):
    
    if not request.user.is_authenticated:
        return Response('User not authenticated', status=401)

    try: 
        userProfile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return Response('Error obteniendo data del userprofile', status = 404)
    

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    mes_resumen =  request.GET.get('current_month', '')
    ws.title = "Resumen del mes " + mes_resumen

    ws.merge_cells("A1:G1")
    ws["A1"] = "Detalles de los movimientos de "+mes_resumen
    ws.merge_cells("A2:G2")
    ws["A2"] = ""
    # Escribir encabezados
    ws.append(["Motivo", "Cantidad", "Fecha", "Estado"])

    # Escribir filas (ejemplo con PaymentForUse)
    try:

        pagos = Payment.objects.filter(order__userProfile=userProfile)

        for pago in pagos:
            ws.append([pago.order.desc, pago.amount, pago.created_at.strftime("%Y-%m-%d"), pago.status])

    except Payment.DoesNotExist:
        return Response('error', status =  400)
    # Respuesta HTTP con el archivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="golocal_detalles{mes_resumen}.xlsx"'
    wb.save(response)
    return response




from decouple import config

APP_ID = config('APP_ID')
APP_SECRET = config('APP_SECRET')
REDIRECT_URI = config('REDIRECT_URI')
MP_ACCESS_TOKEN= config('MP_ACCESS_TOKEN')




import base64
import hashlib
import secrets


@api_view(['GET'])
def generateOauthMpLink(request):
    # ==== CONFIGURA TUS DATOS ==== 
    # ==== GENERAR PKCE ====
    code_verifier = secrets.token_urlsafe(64)
    code_challenge = base64.urlsafe_b64encode(
        hashlib.sha256(code_verifier.encode()).digest()
    ).decode().replace("=", "")

    # ==== GUARDAR PARA USAR EN STEP 2 ====
    with open("pkce_data.txt", "w") as f:
        f.write(code_verifier)

    # ==== CREAR URL DE AUTORIZACIÓN ====
    auth_url = (
        f"https://auth.mercadopago.com/authorization"
        f"?response_type=code"
        f"&client_id=4728062981588764"
        #f"&client_id={210936078469229}" #el d'0avui
        f"&redirect_uri=https://golocalbackend.onrender.com/api/pending/"
        f"&code_challenge={code_challenge}"
        f"&code_challenge_method=S256"
        f"&login_hint=test_user_3270711302169446228@testuser.com"
    )

   
    auth_url += '&state='+ str(request.user.id)
   
    return Response({'link':auth_url}, status = 200)



from cryptography.fernet import Fernet
import os

#url tipo que hit el redirect_url para obtener el accesstoken 
#https://borradoresdeviaje.com/?code=TG-68a4a2e93b2f1a0001aeee5b-2627047012
#http://localhost:8000/api/success_oauth_registration/?code=TG-68a4aa6b7020150001abbc9f-2633951459&state=33


#AQUESTA ES LA QUE M'HA FUNCIONAT
#http://localhost:8000/api/success_oauth_registration/?code=TG-68c094d016c07d0001f8ba11-2633951459&state=3
#http://localhost:8000/api/success_oauth_registration/?code=TG-68c2f3daf5e40700019641e2-2621708065&state=4
#http://localhost:8000/api/success_oauth_registration/?code=TG-68c74b5eec90b30001007b70-2633951459&state=3
import requests

@api_view(['GET'])
def obtainAccessTokenVendedor(request):
    auth_code = request.GET.get('code')
    user_id = request.GET.get('state')

      # viene de la URL
    with open("pkce_data.txt") as f:
        code_verifier = f.read().strip()

    resp = requests.post(
        "https://api.mercadopago.com/oauth/token",
        data={
            "grant_type": "authorization_code",
            "client_id": 4728062981588764,
            "client_secret": 'ct6urSRBJln4XFSznbVm9Y0yAgUuhOBO',
            "code": auth_code,
            "redirect_uri": 'https://golocalbackend.onrender.com/api/pending/',
            "code_verifier": code_verifier
        }

    )

   
    if resp.status_code != 200:
        print("Error OAuth:", resp.text)
        return Response({"error": "No se pudo obtener access token"}, status=400)
    response = resp.json() 

    #response =  {'access_token': 'APP_USR-2369085495887617-081912-ad216cd1cef5503d9bd07a7dff81907a-2633951459', 
    #             'token_type': 'Bearer', 
    #             'expires_in': 15552000, 
    #             'scope': 'offline_access payments read write', 
    #             'user_id': 2633951459, 
    #             'refresh_token': 'TG-68a4aa9309f7650001b985af-2633951459', 
    #             'public_key': 'APP_USR-3917cb81-ca09-4a7b-971f-1d6161e72f56',
    #                'live_mode': True}

    SECRET_KEY = os.environ.get("FERNET_KEY")
    if not SECRET_KEY:
        raise Exception("FERNET_KEY no definida en variables de entorno")
    fernet = Fernet(SECRET_KEY)
    ##{'access_token': 'APP_USR-2369085495887617-081912-ad216cd1cef5503d9bd07a7dff81907a-2633951459', 'token_type': 'Bearer', 'expires_in': 15552000, 'scope': 'offline_access payments read write', 'user_id': 2633951459, 'refresh_token': 'TG-68a4aa9309f7650001b985af-2633951459', 'public_key': 'APP_USR-3917cb81-ca09-4a7b-971f-1d6161e72f56', 'live_mode': True}
    # ===== Cifrar los tokens =====
    encrypted_access = fernet.encrypt(response['access_token'].encode()).decode()
    encrypted_refresh = fernet.encrypt(response['refresh_token'].encode()).decode()


    try:
        userProfile = UserProfile.objects.get(user__id = int(user_id))
        userProfile.mp_access_token = encrypted_access
        userProfile.mp_refresh_token = encrypted_refresh
        userProfile.mp_user_id = response['user_id']
        userProfile.save()
    except UserProfile.DoesNotExist:
        
        with open("mp_tokens_encrypted.txt", "ab") as f:
            # Guardamos cada usuario en una línea
            f.write(b"user_id: " + str(user_id).encode() + b"\n")
            f.write(b"access_token: " + encrypted_access + b"\n")
            f.write(b"refresh_token: " + encrypted_refresh + b"\n")
            f.write(b"mp_user_id: " + response['user_id'] + b"\n\n")
        return ('Error in storing the access token and refresh token')


    return Response({"message": "Tokens obtenidos correctamente"}, status = 200)


@api_view(['GET'])
def createSplitPayment(request):
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)

    try:
        userProfile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response('UserProfile no encontrado', status=400)

    # Descifrar access token del vendedor
    SECRET_KEY = os.environ.get("FERNET_KEY")
    if not SECRET_KEY:
        raise Exception("FERNET_KEY no definida en variables de entorno")
    fernet = Fernet(SECRET_KEY)
    access_token = fernet.decrypt(userProfile.mp_access_token.encode()).decode()
    collector_id = userProfile.mp_user_id  # user_id del vendedor sandbox
    url = "https://api.mercadopago.com/checkout/preferences"

    # Items de prueba
    items = [
        {
            "title": "Producto de prueba",
            "quantity": 1,
            "unit_price": 100.0,
            "currency_id": "ARS"
        }
    ]

    payload = {
        "items": items,
        "back_urls": {
            "success": "https://golocalbackend.onrender.com/api/success/",
            "failure": "https://golocalbackend.onrender.com/api/failure/",
            "pending": "https://golocalbackend.onrender.com/api/pending/"
        },
        "notification_url": "https://golocalbackend.onrender.com/api/pending/",
       # "external_reference": "TEST_PAGO_001",
        "auto_return": "approved",
        "marketplace_fee":10.0,

    }


    sdk = mercadopago.SDK(access_token)
    resp = sdk.preference().create(payload)

    init_point = resp['response']['init_point']
    sandbox_init_point = resp["response"]["sandbox_init_point"]
    return Response({"sandbox_init_point": sandbox_init_point}, status=200)

@api_view(['GET'])
def eventosActivos(request):
    user = request.user

    if not user.is_authenticated:
        return Response('User not authenticated', status=401)

    if not user.profile.creador:
        return Response('No eres creador, no tienes acceso', status=403)
    
    today = timezone.localtime(timezone.now()).replace(hour=0, minute=0, second=0, microsecond=0)
    
    eventos = (
        Activity.objects
        .filter(
            creador=user,
            startDateandtime__gte=today,
            gratis=False,
            control_entradas=True,
            entradas_for_plan__isnull=False
        )
        .order_by('startDateandtime').distinct()
    )
    serializer = ActivitySerializerForGoLocalQR(eventos, many=True)
    return Response(serializer.data, status=200)


@api_view(['POST'])
def createReserva(request):

    user = request.user
    if not user.is_authenticated:
            return Response('User not authenticated', status=401)
    
    try:
        userProfile = UserProfile.objects.get(user = user)
    except UserProfile.DoesNotExist:
        return Response('Error in retrieving UP', status = 400)
    
    try:
        reserva_form = ReservaForm.objects.get(uuid = request.data['uuid'])
    except ReservaForm.DoesNotExist:
        return Response('No se encontro el formulario de reserva', status = 400)

    reserva = Reserva.objects.create(
        reserva_form=reserva_form,
        values=request.data['values'],
    )

    pax = 1 #por defecto reservas para 1
    if 'personas' in request.data['values']:
        pax = request.data['values']['personas']

    reserva_form.confirmados += int(pax)
    reserva_form.save()
    
    if reserva_form.activity != None:
        userProfile.activities.add(reserva_form.activity)
    else:
        userProfile.promos.add(reserva_form.promo)

    userProfile.save()
    email = EmailMessage(
        subject="Tu reserva para el evento 🎟️",
        body=f"Hola {userProfile.user.username}, aquí esta tu confirmación de reserva para {reserva_form.activity.name}.\n¡Nos vemos el {reserva_form.activity.startDateandtime}!\n\n\n\n :) ",
        from_email="no-reply@miapp.com",
        to=[user.email],
    )
    email.send()

    serializer = None
    if reserva_form.activity != None:
        asistentes = UserProfile.objects.filter(activities=reserva_form.activity)
        serializer = UserProfileBasicSerializer(asistentes, many=True, context={"request": request})

    elif reserva_form.promo != None:
        asistentes = UserProfile.objects.filter(promos=reserva_form.promo)
        serializer = UserProfileBasicSerializer(asistentes, many=True, context={"request": request})

    return Response(serializer.data, status=200)

@api_view(['POST', 'GET'])
def successMP(request):
    return Response({"message": "Success"}, status=200)

@api_view(['POST', 'GET'])
def failureMP(request):
    return Response({"message": "Failure"}, status=200)

@api_view(['POST', 'GET'])
def pendingMP(request):
    return Response({"message": "Pending"}, status=200)

import warnings
import mercadopago
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@api_view(['POST'])
def createCompraSimple(request):
    print('crete compra simple: ', request.data)
    user = request.user
    if not user.is_authenticated:
        return Response('User not authenticated', status=401)   

    try:
        userProfile = UserProfile.objects.get(user=user, creador = True)
    except UserProfile.DoesNotExist:
        return Response('UserProfile no encontrado o el usuario no es creador', status = 400)

    data = request.data
    #data['type'] pot ser: 
    ##### --> 'compra_bono'
    ##### --> 'extend_rango'
    ##### --> 'compra_tickets'

    if data['type'] == 'compra_bono':
        print('entra al compra bono!')
       
        try:
            bono = Bono.objects.get(uuid = data['uuid'])
            seller = UserProfile.objects.get(id=1)
            payment = Payment.objects.create(
                userProfile = userProfile,
                status = 'pending',
                descripcion = bono.name,
                event_type = 'interno',
                total_amount = bono.price,
                platform_fee = 0.0,
                seller_amount = bono.price,
                seller = seller, #jo
                activity = None,
            )
            metadata = {"type": "compra-bono", "id": bono.id}
            successUrl = "https://go-local-web.vercel.app/"
        except Bono.DoesNotExist:
            print('error que no hi ha bono')
            return Response('Bono no encontrado', status=404)
       

    elif data['type'] == 'extend_rango':
        try:
            rango = PaymentEventsRanges.objects.get(uuid = data['uuid'])  

            seller = UserProfile.objects.get(id=1)
            payment = Payment.objects.create(
                userProfile = userProfile,
                status = 'pending',
                descripcion = rango.name,
                event_type = 'interno',
                total_amount = rango.price-userProfile.payment_events_range.price,
                platform_fee = 0.0,
                seller_amount = rango.price-userProfile.payment_events_range.price,
                seller = seller, #jo
                activity = None,
            )
            successUrl = "https://go-local-web.vercel.app/"
        except PaymentEventsRanges.DoesNotExist:
            return Response('Rango no encontrado', status=404)
        metadata = {"type": "extend-rango", "id": rango.id}
    elif data['type'] == 'compra-tickets':
       
        #request['entradas'] = [{'uuid':'xxx', 'amount':2},{'uuid':'yyy','amount':1}] --> 2 normales y 1 VIP i.e.

        platform_fee = 0
        seller_amount = 0
        seller = None
        activity = None
        for entrada in data['entradas']:
            if entrada['amount'] > 0:
                try:
                    entrada_comprar = EntradasForPlan.objects.get(uuid=entrada['uuid'])
                    seller = entrada_comprar.activity.creador
                    activity = entrada_comprar.activity
                 
                    seller_amount += Decimal(entrada_comprar.precio) * Decimal(entrada['amount'])
                   
                    platform_fee +=  (Decimal(seller.platform_fee) / Decimal(100))*Decimal(entrada_comprar.precio)*Decimal(entrada['amount'])
                  
                except EntradasForPlan.DoesNotExist:
                    return Response('No encontramos las entradas', status =400)
            
     
        
        existing_payment = Payment.objects.filter(
            userProfile=userProfile,
            status='pending',
            activity=activity
        ).first()

        
        if existing_payment:
            payment = existing_payment
        else: 
            try: 
                payment = Payment.objects.create(
                    userProfile = userProfile,
                    status = 'pending',
                    descripcion = 'Compra de tickets',
                    event_type = 'externo',
                    total_amount = platform_fee+seller_amount,
                    platform_fee = platform_fee,
                    seller_amount = seller_amount,
                    seller = seller,
                    activity = activity,
                )
            except IntegrityError as e:
                return Response(f'Error creando el payment object: {str(e)}', status=400)
            except Exception as e:
                return Response(f'Error inesperado al crear payment: {str(e)}', status=500)
            
        metadata = {"type": "compra-tickets", "entradas": [{"uuid": e["uuid"], "amount": e["amount"]}for e in data["entradas"] if e["amount"] > 0],"name":data['name'], "email":data['email']}
        successUrl = "golocal://activity/"+str(activity.uuid)

   
    print('arriba a crear la preferència')
    sdk = mercadopago.SDK(os.environ.get("MP_ACCESS_TOKEN"))

    preference_data = {
    "items": [
        {
            "title": payment.descripcion,
            "quantity": 1,
            "unit_price": float(payment.total_amount),
        }
    ],
    "payer": {
        "email": userProfile.user.email
    },
     "back_urls": {
        "success": successUrl,
        "failure": "https://borradoresdeviaje.com/series/series-australia.html",
        "pending": "https://borradoresdeviaje.com/series/series-tailandia.html"
    },
    "notification_url": "https://golocalbackend.onrender.com/api/webhook_mp/",

    "external_reference": "order_"+str(payment.uuid),

    # 👇 datos adicionales (opcionales)
    "metadata": metadata,
    }

    preference_response = sdk.preference().create(preference_data)
    print('preference response: ', preference_response)
    warnings.filterwarnings("ignore", category=UserWarning)


    preference_response = sdk.preference().create(preference_data)
    preference = preference_response.get("response")
    if preference and "init_point" in preference:
        return Response({"init_point": preference["init_point"]}, status=200)
    else:
        return Response({"error": "Failed to create preference"}, status=400)



@csrf_exempt
@api_view(['POST'])
def webhook_mp(request):
    body = json.loads(request.body.decode("utf-8"))
    
    # Caso 1: webhook tipo payment (más confiable)
    payment_id = None
    if "data" in body and "id" in body["data"]:
        payment_id = body["data"]["id"]
    elif "resource" in body and body.get("topic") == "payment":
        payment_id = body["resource"]

    sdk = mercadopago.SDK(os.environ.get("MP_ACCESS_TOKEN"))

    if payment_id:
        payment_response = sdk.payment().get(payment_id)
        payment = payment_response["response"]
        external_reference = payment.get("external_reference")
        payment_uuid = external_reference.split("_")[1]
        metadata = payment.get("metadata", {})
        status = payment.get("status")
        amount = payment.get("transaction_amount")

        try:
            payment = Payment.objects.get(uuid = payment_uuid)
        except Payment.DoesNotExist:
            return Response('Error, payment not found', status= 400)
        
        payment.metadata = metadata
        payment.payout_status = 'pending'
        payment.status = status
        payment.save()

        ####SE COMPRO UN BONO DE PLANES GRATIS
        if metadata['type'] == 'compra-bono' and status == 'approved' and not payment.applied:
            try:
                bono = Bono.objects.get(id = metadata['id'])
                userProfile = payment.userProfile
                userProfile.available_planes_gratis += bono.amount
                userProfile.save()
                payment.applied = True
                payment.save()
            except Bono.DoesNotExist:
                pass
        ####SE EXTENDIÓ EL RANGO DE PLANES PAGOS
        elif metadata['type'] == 'extend-rango' and status == 'approved' and not payment.applied:
            try:
                rango = PaymentEventsRanges.objects.get(id = metadata['id'])
                userProfile = payment.userProfile
                userProfile.payment_events_range = rango
                userProfile.save()
                payment.applied = True
                payment.save()
            except PaymentEventsRanges.DoesNotExist:
                pass
        elif metadata['type'] == 'compra-tickets' and status == 'approved' and not payment.applied:
            tickets = []
            oneTime = True
            for e in metadata['entradas']:
                try:
                    entrada = EntradasForPlan.objects.get(uuid = e['uuid'])
                    for _ in range(e['amount']):
                        ticket = Ticket.objects.create(
                            user_profile=payment.userProfile,
                            entrada=entrada,
                            nombre= metadata['name'],
                            email= metadata['email'],
                            fecha_compra=timezone.now(),
                            precio = entrada.precio
                        )
                        entrada.disponibles = entrada.disponibles -1
                        tickets.append(generar_ticket_pdf(metadata['name'],entrada.activity.name, entrada.activity.startDateandtime,ticket.qr_code)) #,os.path.join(settings.BASE_DIR, 'api', 'assets', 'backgroundticketsimage.png')))
                        
                    if oneTime: 
                        payment.userProfile.activities.add(entrada.activity)
                        oneTime = False
                    entrada.save()


                except EntradasForPlan.DoesNotExist:
                    return Response('Error no such Entrada encontrada', status = 400)
            payment.applied = True
            payment.save()
            
            email = EmailMessage(
                subject="Tu ticket para el evento 🎟️",
                body=f"Hola {metadata['name']}, adjuntamos tu ticket para {entrada.activity.name}.\n¡Nos vemos el {entrada.activity.startDateandtime}!",
                from_email="golocalgolocal@gmail.com",
                to=[metadata['email']],
            )


            # Adjuntar PDF
            for i, ticket in enumerate(tickets):
            # ticket es un BytesIO o archivo
                ticket.seek(0)
                email.attach(f"ticket_{i+1}.pdf", ticket.read(), "application/pdf")
        #  email.attach("ticket.pdf", tickets.read(), "application/pdf")

            # Enviar
            email.send()

        return Response({"status": "processed"})

    # Caso 2: webhook tipo merchant_order → podés ignorar o usar si querés
    if body.get("topic") == "merchant_order":
        return Response({"status": "merchant_order_ignored"})

    return Response({"status": "ok"})






    

